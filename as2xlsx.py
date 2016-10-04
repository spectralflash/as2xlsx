# -*- coding: utf-8 -*-
import os
import re
import sys
import mmap
import xlsxwriter
import openpyxl
import datetime
import codecs
import contextlib
from argparse import ArgumentParser
from numbers import Number

__author__ = 'Andrey Berestyansky'


def get_params():

    parser = ArgumentParser("Parses .as files from a specified folder and all subfolders. Generates XLSX file "
                            "with two columns: unique ID (file name + offset), Chinese text (all quoted substrings "
                            "containing at least one non-ASCII character)")
    parser.add_argument("folder", help="folder with .as files")
    parser.add_argument("xlsx", nargs="?", help="XLSX file with translation to insert into .as files; "
                                     "translation must be in column 3; if omitted, a new XLSX file will be generated")
    return parser.parse_args()


def is_ascii(s):
    return all(ord(c) < 128 for c in s)


def timestamp():
    return datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def find_matches(path, file_extensions, sequence):
    """
    :param content_folder: folder to look up, string
    :param sequence: compiled regex to look for, re.compile(r'([\"\'])(.+?)\1', re.MULTILINE)
    :param file_extensions: file extensions, string or tuple of strings: '.as' or ('.as', '.txt')
    :return: dict of dicts: {string ID: {'zh': 'match'}}, where string ID is filename + # + offset (in characters)
    """
    match_dict = {}
    for folder, _, files in os.walk(path):
        for filename in files:
            if not filename.endswith(file_extensions):
                continue
            with codecs.open(os.path.join(folder, filename), 'r', 'utf8') as as_file:
                with contextlib.closing(mmap.mmap(as_file.fileno(), 0, access=mmap.ACCESS_READ)) as mapped_file:
                    data = mapped_file.read(-1)
                    try:
                        data = data.decode('utf8')
                    except Exception, e:
                        data = data.decode('windows-1252')

            for match in sequence.finditer(data):
                if is_ascii(match.group()):
                    continue
                key = os.path.join(folder.replace(path, "")[1:], filename) + "#" + str(match.start()).zfill(12)
                value = match.group()[1:-1]
                match_dict[key] = {'zh': value}
    return match_dict


def save_xlsx(content, name):
    """
    :param content: dict of dicts, such as: {'string ID':{'ru':'Russian string', 'en':'English string'}}
    :param name: 'filename.xlsx'
    """

    print "Strings saved to XLSX:", len(content)
    wb = xlsxwriter.Workbook(name)
    ws = wb.add_worksheet()
    ws.title = "Data"
    ws.write(0, 0, 'ID')

    # write headers, such as: 'ru', 'en'
    langs = list(content.values()[0].keys())
    for lang in langs:
        ws.write(0, langs.index(lang) + 1, lang)

    str_num = 1
    for string_id in sorted(content.keys()):
        entry = content.pop(string_id)
        # write string ID
        ws.write(str_num, 0, string_id)
        # write string content
        for lang in langs:
            ws.write(str_num, langs.index(lang) + 1, entry[lang])
        str_num += 1

    wb.close()


def dict_from_xlsx(filename):
    """
    :param filename: 'filename.xlsx': first string: 'path', 'ru', 'en' etc.; first column: string IDs
    :return: dict of dicts, such as: {'string ID':{'ru':'Russian string', 'en':'English string'}}
    """

    wb = openpyxl.load_workbook(filename=filename, data_only=True, read_only=True)
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])

    headers = []

    for row in ws.get_squared_range(1, 1, ws.max_column, 1):
        for cell in row:
            headers.append(cell.value)

    res_dict = {}  # defaultdict(dict)
    for row in ws.get_squared_range(1, 2, ws.max_column, ws.max_row):
        str_id = str(row[0].value)
        res_dict[str_id] = {}  # defaultdict(dict)
        for cell in row[1:]:
            if cell.value is None:
                res_dict[str_id][headers[row.index(cell)]] = ""
            elif isinstance(cell.value, Number):
                res_dict[str_id][headers[row.index(cell)]] = str(cell.value)
            else:
                res_dict[str_id][headers[row.index(cell)]] = cell.value

    return res_dict


def single_lang_from_dict(source_dict, target_lang):
    """
    :param source_dict: dict of dicts, multiple languages: {'string ID':{'ru':'Russian string', 'en':'English string'}}
    :param target_lang: string, such as 'ru' or 'en'
    :return: dict of dicts, single language: {'string ID':{'ru':'Russian string'}}
    """
    return {string_id: {target_lang: source_dict[string_id][target_lang]} for string_id in source_dict}


def get_target_lang(filename):
    wb = openpyxl.load_workbook(filename=filename, data_only=True, read_only=True)
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    return ws.cell(None, 1, 3).value


if __name__ == '__main__':

    seq = re.compile(r'([\"\'])(.+?)\1', re.MULTILINE)

    args = get_params()

    content_folder = os.path.normpath(args.folder)
    os.chdir(os.path.dirname(content_folder))

    dict_as = find_matches(content_folder, '.as', seq)

    if args.xlsx:

        target_lang = get_target_lang(args.xlsx)
        dict_xlsx = dict_from_xlsx(args.xlsx)

        # integrity check
        dict_xlsx_source = single_lang_from_dict(dict_xlsx, 'zh')
        # if dict_xlsx_source != dict_as:
        #     sys.exit("String IDs in the XLSX file do not match the .as files!")

        dict_xlsx_target = single_lang_from_dict(dict_xlsx, target_lang)
        sorted_keys = sorted([key for key in dict_xlsx_target.keys()])

        offset_correction={}
        for string_id in dict_xlsx_target:
            offset_correction[string_id.split("#")[0]]=0

        for string_id in sorted_keys:
            print string_id.split("#")[0]
            with open(os.path.join(content_folder, string_id.split("#")[0]), 'r+b') as as_file:
                with contextlib.closing(mmap.mmap(as_file.fileno(), 0, access=mmap.ACCESS_WRITE)) as mapped_file:
                    current_data = mapped_file.read(-1).decode('utf8')

                    start_pos = int(string_id.split("#")[1]) + offset_correction[string_id.split("#")[0]]
                    end_pos = int(string_id.split("#")[1]) + len(dict_xlsx_source[string_id]['zh']) + \
                              offset_correction[string_id.split("#")[0]]
                    en_length = len(dict_xlsx_target[string_id][target_lang])
                    zh_length = len(dict_xlsx_source[string_id]['zh'])

                    offset_correction[string_id.split("#")[0]] = \
                        offset_correction[string_id.split("#")[0]] - zh_length + en_length

                    new_data = current_data[:start_pos + 1] + dict_xlsx_target[string_id][target_lang] \
                               + current_data[end_pos + 1:]
                    new_data = new_data.encode('utf-8')
                    mapped_file.resize(len(new_data))

                    # overwrite with new_data
                    mapped_file.seek(0)
                    mapped_file.write(new_data)
                    mapped_file.flush()

    else:
        save_xlsx(dict_as, "as2xlsx_" + timestamp() + ".xlsx")

